import pprint
import openpyxl
import requests

# 读取excel数据
def ReadDate(filename,sheetname):
    #获取工作簿
    wk= openpyxl.open(filename=filename)
    #获取工作表
    sheet1=wk[sheetname]
    datalist = []
    #如何取得excel所有的数据
    for x in range(2,sheet1.max_row+1):
        # 把用例数据存储到字典
        case = dict(case_id=sheet1.cell(x,1).value,
                    url=sheet1.cell(row=x, column=5).value,
                    data=sheet1.cell(row=x, column=6).value,
                    excepted=sheet1.cell(row=x, column=7).value)
        # print(case)
        datalist.append(case)
    return datalist

def ReadDate_token(filename,sheetname):
    #获取工作簿
    wk= openpyxl.open(filename=filename)
    #获取工作表
    sheet1=wk[sheetname]
    datalist = []
    #如何取得excel所有的数据
    for x in range(2,sheet1.max_row+1):
        # 把用例数据存储到字典
        case = dict(case_id=sheet1.cell(x,1).value,
                    url=sheet1.cell(row=x, column=5).value,
                    pwd=sheet1.cell(row=x, column=6).value,
                    data=sheet1.cell(row=x, column=7).value,
                    excepted=sheet1.cell(row=x, column=8).value)
        # print(case)
        datalist.append(case)
    return datalist

#发送请求
def api_request(url,json):
    headers={"X-Lemonban-Media-Type":"lemonban.v2",
    "Content-Type":"application/json"}
    response = requests.post(url=url, json=json, headers=headers)
    return response.json()

def api_request_token(url,json,token):
    headers={"X-Lemonban-Media-Type":"lemonban.v2",
    "Content-Type":"application/json"
    }
    headers["Authorization"] = "Bearer "+token
    response = requests.post(url=url, json=json, headers=headers)
    return response.json()

# 登录
def login(json):
    url="http://api.lemonban.com/futureloan/member/login"
    headers = {"X-Lemonban-Media-Type": "lemonban.v2",
               "Content-Type": "application/json"}
    response = requests.post(url=url, json=json, headers=headers)
    return response.json()

# 回写测试执行的结果
def write_result(filename,sheetname,row,column,result):
    wk = openpyxl.open(filename=filename)
    sheet1 = wk[sheetname]
    sheet1.cell(row,column).value = result
    wk.save(filename)

def exec_case(filename,sheetname):
    # 获取excel所有用例的数据
    cases = ReadDate(filename,sheetname)

    #遍历数据，执行请求
    for case in cases:
        case_id = case["case_id"]
        url = case["url"]
        # 去引号把data数据转换为字典
        data = eval(case["data"])

        # 发送请求获取响应的数据
        response = api_request(url,data)
        print(response)
        # 获取期望结果
        excepted_msg = eval(case["excepted"])["msg"]
        print(f"预期结果为：{excepted_msg}")

        # 获取期望结果
        real_msg = response["msg"]
        print(f'用例实际结果:', real_msg)

        #结果比对
        if real_msg == excepted_msg:
            print(f"用例{case_id}测试执行通过")
            write_result(filename=filename, sheetname=sheetname,
                         row=case_id + 1, column=8, result='通过')
        else:
            print(f"用例{case_id}测试执行不通过")
            write_result(filename=filename, sheetname=sheetname,
                         row=case_id + 1, column=8, result='不通过')
        print('*' * 30)

def exec_case_token(filename,sheetname):
    # 获取excel所有用例的数据
    cases = ReadDate_token(filename, sheetname)

    # 遍历数据，执行请求
    for case in cases:
        case_id = case["case_id"]
        url = case["url"]

        # 去引号把data数据转换为字典
        data = eval(case["data"])
        pwd = eval(case["pwd"])

        # 发送请求获取响应的数据
        response = login(pwd)
        # print(response)
        token = response["data"]["token_info"]["token"]
        response = api_request_token(url, data,token)

        # 获取期望结果
        excepted_msg = eval(case["excepted"])["msg"]
        print(f"预期结果为：{excepted_msg}")

        # 获取期望结果
        real_msg = response["msg"]
        print(f'用例实际结果:', real_msg)

        # 结果比对
        if real_msg == excepted_msg:
            print(f"用例{case_id}测试执行通过")
            write_result(filename=filename, sheetname=sheetname,
                         row=case_id + 1, column=8, result='通过')
        else:
            print(f"用例{case_id}测试执行不通过")
            write_result(filename=filename, sheetname=sheetname,
                         row=case_id + 1, column=8, result='不通过')
        print('*' * 30)

# 执行注册用例
exec_case('test_case_api.xlsx',"register")
# 执行登录用例
# exec_case('test_case_api.xlsx',"login")
#执行投资用例
# exec_case_token('test_case_api.xlsx',"invest")
